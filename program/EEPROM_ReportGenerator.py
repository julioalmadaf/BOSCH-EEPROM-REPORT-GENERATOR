##################
#### Libraries ###
#### Librerias ###
##################
import os
import sys
import tkinter as tk
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as e

from os import listdir
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter import *

try:
        import xlrd
except ImportError:
        os.system('python -m pip install xlrd')

try:
        import xlwt
except ImportError:
        os.system('python -m pip install xlwt')

try:
        from pathlib import Path
except ImportError:
        os.system('python -m pip install pathlib')

try:
        import shutil
except ImportError:
        os.system('python -m pip install pytest-shutil')

try:
        from PIL import ImageTk, Image
except ImportError:
        os.system('python -m pip install pillow')

try:
        from openpyxl import load_workbook
except ImportError:
        os.system('python -m pip install openpyxl')

try:
        import pandas as pd
except ImportError:
        os.system('python -m pip install pandas')

try:
        import win32com.client
except ImportError:
        os.system('python -m pip install pypiwin32')
        os.system('python -m pip install pywin32')

###########################
#### Global Variables ###
#### Variables globales ###
###########################
rutaCNT = "0"
rutaArchivoCNT = "0"
archivoCNTCargado = 0
rutaReportePrevio = "0"
archivoReportePrevioCargado = 0
BBNumber = 0
Baseline = 0

estadoCheckButton = 0
estadoEtiqueta = 1

################
#### Events ###
### Eventos ###
################
def newProgram():

        global rutaCNT
        global rutaArchivoCNT
        global archivoCNTCargado
        global rutaReportePrevio
        global archivoReportePrevioCargado
        global estadoCheckButton
        global estadoEtiqueta

        #Hide Buttons that are not being used
        #Ocultar botones innecesarios.
        button_PreviousReport.grid_remove()
        enable_button.grid_remove()
        button_GenerateReport.grid_remove()

        button_CNT.configure(style='button_style1.TButton')
        enable_button.deselect()
        button_PreviousReport.configure(style='button_style2.TButton')
        button_GenerateReport.configure(style='button_style2.TButton')

        #Previous report button disabled
        #Comenzar con el boton de reporte previo deshabilitado.
        estadoCheckButton = 0
        button_PreviousReport.state(["disabled"])

        label.configure(text="EEPROM report generator")
        estadoEtiqueta = 1	
        #Message EEPROM report generator displayed
        #La etiqueta muestra el primer mensaje (EEPROM report generator)

        rutaCNT = "0"
        rutaArchivoCNT = "0"
        archivoCNTCargado = 0
        rutaReportePrevio = "0"
        archivoReportePrevioCargado = 0
        BBNumber = 0
        Baseline = 0
        estadoCheckButton = 0
        estadoEtiqueta = 1

def exitProgram():
        
        global estadoEtiqueta

        label.configure(text="Closing program")

        #Ask if the user wants to exit the program
        #Preguntar al usuario si desea salir del programa.
        salir = messagebox.askyesno(message="Do you want to close the program?", title="Close program")

        if(salir == 1):
                sys.exit(0)
        
        if(estadoEtiqueta == 1):
        	label.configure(text="EEPROM report generator")
        elif(estadoEtiqueta == 2):
        	label.configure(text="CNT file loaded")
        elif(estadoEtiqueta == 3):
        	label.configure(text="CNT file loaded - Previous Report loaded")
        else:	#Si entra aqui es porque algo grave paso/ Error happened
        	label.configure(text="EEPROM report generator")

def aboutProgram():
        
        global estadoEtiqueta

        label.configure(text="About the program")

        messagebox.showinfo("About EEPROM report generator", "This software has been released by Julio Cesar Almada Fuerte and Ruben Barajas Curiel")

        if(estadoEtiqueta == 1):
        	label.configure(text="EEPROM report generator")
        elif(estadoEtiqueta == 2):
        	label.configure(text="CNT file loaded")
        elif(estadoEtiqueta == 3):
        	label.configure(text="CNT file loaded - Previous Report loaded")
        else:	#Si entra aqui es porque algo grave paso/Error happened
        	label.configure(text="EEPROM report generator")

def helpProgram():
        
        global estadoEtiqueta

        label.configure(text="Help")

        messagebox.showinfo("Help", "Visit the following link to get more information about this software")

        if(estadoEtiqueta == 1):
        	label.configure(text="EEPROM report generator")
        elif(estadoEtiqueta == 2):
        	label.configure(text="CNT file loaded")
        elif(estadoEtiqueta == 3):
        	label.configure(text="CNT file loaded - Previous Report loaded")
        else:	#Si entra aqui es porque algo grave paso/Error happened
        	label.configure(text="EEPROM report generator")

def cntButton():

    global rutaCNT
    global archivoCNTCargado
    global estadoEtiqueta
    global rutaReportePrevio
    global archivoReportePrevioCargado
    global estadoCheckButton

    #Open Dialog box to search for a file
    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening CNT file")
    rutaCNT = filedialog.askopenfilename(filetypes = (("All CNT files","*.cnt"),("All files","*.*")))

    #Verify that the file selected is a CNT file or that a file was selected
    #Verificar que sea archivo CNT o que se haya agregado un archivo.
    if(rutaCNT.find(".cnt") == -1):         #No se selecciono un archivo CNT/No CNT file selected
        
        messagebox.showerror("Error", "Not .cnt file selected")

        #Hide unnecesary buttons.
        #Ocultar botones innecesarios.
        button_PreviousReport.grid_remove()
        enable_button.grid_remove()
        button_GenerateReport.grid_remove()

        button_CNT.configure(style='button_style1.TButton')
        enable_button.deselect()
        button_PreviousReport.configure(style='button_style2.TButton')
        button_GenerateReport.configure(style='button_style2.TButton')

        #Previous report button disabled
        #Comenzar con el boton de reporte previo deshabilitado.
        estadoCheckButton = 0
        button_PreviousReport.state(["disabled"])

        label.configure(text="EEPROM report generator")
        estadoEtiqueta = 1	#La etiqueta muestra el primer mensaje (EEPROM report generator)/Message displayed(EEPROM report generator)

        rutaCNT = "0"
        rutaArchivoCNT = "0"
        archivoCNTCargado = 0
        rutaReportePrevio = "0"
        archivoReportePrevioCargado = 0
        BBNumber = 0
        Baseline = 0
        estadoCheckButton = 0
        estadoEtiqueta = 1

    elif(rutaCNT == ""):                    #Se dio al boton cancelar./Click on Cancel button
		#Ocultar botones innecesarios.
        button_PreviousReport.grid_remove()
        enable_button.grid_remove()
        button_GenerateReport.grid_remove()

        button_CNT.configure(style='button_style1.TButton')
        enable_button.deselect()
        button_PreviousReport.configure(style='button_style2.TButton')
        button_GenerateReport.configure(style='button_style2.TButton')

        #Comenzar con el boton de reporte previo deshabilitado./Previous report button disabled
        estadoCheckButton = 0
        button_PreviousReport.state(["disabled"])

        label.configure(text="EEPROM report generator")
        estadoEtiqueta = 1	#La etiqueta muestra el primer mensaje (EEPROM report generator)/essage displayed(EEPROM report generator)

        rutaCNT = "0"
        rutaArchivoCNT = "0"
        archivoCNTCargado = 0
        rutaReportePrevio = "0"
        archivoReportePrevioCargado = 0
        BBNumber = 0
        Baseline = 0
        estadoCheckButton = 0
        estadoEtiqueta = 1

    else:                                   #Se selecciono el archivo correctamente./File selected correctly
        archivoCNTCargado = 1
        button_CNT.configure(style='button_style2.TButton')
        enable_button.grid()            #Ahora se puede mostrar el checkbuton para habilitar el boton de reporte previo./Checkbutton displayed for previous report
        button_PreviousReport.state(["disabled"])
        button_PreviousReport.grid()    #Ahora se puede mostrar el boton de reporte previo como opcional./Previous report button shown as optional(can be toggled)
        button_GenerateReport.grid()    #Ahora se puede mostrar el boton de generar reporte./ Generate report button displayed
        button_GenerateReport.configure(style='button_style1.TButton')
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)/ CNT file loaded message displayed

def enableButtonRP():

        global estadoCheckButton
        global archivoReportePrevioCargado

        estadoCheckButton = estadoCheckButton ^ 1

        #Change the state of the previous report depending on the checkbutton
        #Cambiar estado del boton de reporte previo dependiendo del CheckButton.
        if(estadoCheckButton == 0):
                button_PreviousReport.state(["disabled"])
        else:
                button_PreviousReport.state(["!disabled"])

        #If a previous report is loaded, show or hide message
        #Si ya existe un reporte cargado, esconder o mostrar el label actualizado con la accion a hacer.
        if(archivoReportePrevioCargado == 1 and estadoCheckButton == 1):
        	label.configure(text="CNT file loaded - Previous Report loaded")
        	estadoEtiqueta = 3	#La etiqueta muestra el tercer mensaje (CNT file loaded - Previous Report loaded)/(CNT file loaded - Previous Report loaded) message shown
        if(archivoReportePrevioCargado == 1 and estadoCheckButton == 0):
        	label.configure(text="CNT file loaded")
        	estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)/ (CNT file loaded) message shown

def previousReport():

    global rutaReportePrevio
    global archivoReportePrevioCargado
    global estadoEtiqueta

    #Open dialog box to search for a previous report
    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening previous report file")
    rutaReportePrevio = filedialog.askopenfilename(filetypes = (("All Excel files","*.xlsx"),("All files","*.*")))

    #Verify that the file is XLSX or a file was added
    #Verificar que sea archivo XLSX o que se haya agregado un archivo.
    if(rutaReportePrevio.find(".xlsx") == -1):       #No se selecciono un archivo XLXS/No xlxs file selected
        messagebox.showerror("Error", "Not .xlsx file selected")
        archivoReportePrevioCargado = 0
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)/(CNT file loaded) message shown
    elif(rutaReportePrevio== ""):                    #Se dio al boton cancelar./Cancel button clicked
        messagebox.showerror("Error", "Not .xlsx file selected")
        archivoReportePrevioCargado = 0
        label.configure(text="CNT file loaded")
        estadoEtiqueta = 2	#La etiqueta muestra el segundo mensaje (CNT file loaded)/(CNT file loaded) message shown
    else:                                   #Se selecciono el archivo correctamente./ File selected correctly
        archivoReportePrevioCargado = 1
        label.configure(text="CNT file loaded - Previous Report loaded")
        estadoEtiqueta = 3	#La etiqueta muestra el tercer mensaje (CNT file loaded - Previous Report loaded)/(CNT file loaded - Previous Report loaded) message shown

def createReport():

    global archivoCNTCargado
    global archivoReportePrevioCargado
    global rutaCNT
    global rutaArchivoCNT
    global rutaReportePrevio
    global BBNumber
    global Baseline
    global estadoEtiqueta
    global estadoCheckButton

    label.configure(text="Creating report")

    #Garantizar que se haya seleccionado un archivo CNT./Make sure a CNT file was selected
    if(archivoCNTCargado == 1):                 
        archivoCNTCargado = 0

        #Extraer BBNumber y Baseline./Obtain BBNumber and Baseline
        BBNumber = rutaCNT.split('_')[3]
        Baseline = rutaCNT.split('_')[4]

        #Crear archivo Excel./ Create Excel file
        rutaArchivoCNT = os.path.dirname(rutaCNT)
        shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/fillexcel.xlsx")

        #Rellena Excel/Fill excel
        fillExcel()

        messagebox.showinfo("Report created", "Report created successfully")

    else:
        messagebox.showerror("Error", "Not .cnt file selected")

    #Ocultar botones innecesarios./Hide unnecesary buttons
    button_PreviousReport.grid_remove()
    enable_button.grid_remove()
    button_GenerateReport.grid_remove()

    button_CNT.configure(style='button_style1.TButton')
    enable_button.deselect()
    button_PreviousReport.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style2.TButton')

    #Comenzar con el boton de reporte previo deshabilitado./Start with the previous report button hidden
    estadoCheckButton = 0
    button_PreviousReport.state(["disabled"])

    label.configure(text="EEPROM report generator")
    estadoEtiqueta = 1	#La etiqueta muestra el primer mensaje (EEPROM report generator)/(EEPROM report generator) message shown

    rutaCNT = "0"
    rutaArchivoCNT = "0"
    archivoCNTCargado = 0
    rutaReportePrevio = "0"
    archivoReportePrevioCargado = 0
    BBNumber = 0
    Baseline = 0
    estadoCheckButton = 0
    estadoEtiqueta = 1

################
#### Objetos ###
#### Objects ###
################
root = Tk()

rutaActual = os.getcwd()
img = ImageTk.PhotoImage(Image.open(rutaActual + "/bosch.png"))

panelElements = ttk.Frame(root, padding=(3,3,12,12))
panelImage = ttk.Frame(panelElements, borderwidth=5, relief="sunken", width=200, height=200)

label = ttk.Label(panelElements, text="EEPROM report generator", font=("Tahoma", 25, 'bold'))
button_CNT = ttk.Button(panelElements, text="Select CNT file", style="TButton", command=cntButton)
button_PreviousReport = ttk.Button(panelElements, text="Select previous report", style="TButton", command=previousReport)
button_GenerateReport = ttk.Button(panelElements, text="Generate report", style="TButton", command=createReport)
enable_button = Checkbutton(panelElements, text="Enable previous report button", onvalue=1,offvalue=0, command=enableButtonRP)
image = tk.Label(panelImage, image = img)
menubar = Menu(panelImage)

button_style1 = ttk.Style()
button_style2 = ttk.Style()

##################
#### Funciones ###
#### Functions ###
##################
def fillExcel():

    global rutaArchivoCNT
    global estadoCheckButton
    
    #Open log file.
    logProgram = open(rutaArchivoCNT + "/logProgram" + str(BBNumber) + ".txt","w+")

    logProgram.write("Generating new Excel file\r\n")

    #Create new excel file
    #Genera archivo excel
    wb = load_workbook(filename = rutaArchivoCNT + "/fillexcel.xlsx")
    ws = wb.active
    
    #Reads XML file(CNT)
    #Lee archivo XML.
    tree = ET.parse(rutaCNT)
    logProgram.write("CNT file read.\r\n\r\n")

    #Gets root from XML
    #Obtiene el root del XML.
    root = tree.getroot()
    
    #counter to add elements to the excel file
    #Counter para ir agregando elementos en excel.
    CounterFilasExcel = 11
    CounterAll=0
    DeleteRepeat=0
    NoRepeat=0
    
    ######################################################
    #Guarda el nombre del proyecto.
    #Saves project name
    for project in root.iter('PROJECT-INFO'):
        logProgram.write("PROJECT-INFO\r\n")
        PD = project.find('PROJECT-DESC')
        if PD is not None:
                logProgram.write("      PROJECT-DESC -- " + PD.text + "\r\n")

    ######################################################
    #Guarda el nombre del responsable.
    #Saves responsible name
    for info in root.iter('RESPONSIBLE'):
        logProgram.write("RESPONSIBLE \r\n")
        PN = info.find('PERSON-NAME')
        if PN is not None:
            logProgram.write("      PERSON-NAME -- " + PN.text + "\r\n")

    ######################################################
    #Busca el nodo sesion en todo el arbol.
    #Searches for session node on all the tree structure
    logProgram.write("\r\nAdding datapointers to Excel file according to session.\r\n\r\n")
    logProgram.write("SESSIONS\r\n")
    for session in root.iter('SESSION'):
        #Looks for the type of sessions and their names
        #Busca en los tipos de sesiones que nombre tiene.
        sessionN = session.find('SESSION-NAME')
        if sessionN is not None:
                logProgram.write("      SESSION-NAME -- " + sessionN.text + "\r\n")
                #ALL session
                #Cuando la sesion es ALL.
                if(sessionN.text == '__ALL__'):
                        #To keep track of excel lines order
                        #Para no alterar el orden de las filas del excel.
                        tempCounter=CounterFilasExcel
                        #Obtains datapointer name from the item
                        #Obtiene el Datapointer-name del item.
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text    #Guarda el valor en el excel./Saves the value on the excel file
                                CounterAll+=1
                        
                        tempCounter = CounterFilasExcel
                        #Obtains the datapointer-ident from the item
                        #Obtiene el Datapointer-ident del item.
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                tempCounter += 1
                                logProgram.write("                      "+ws['A'+ str(tempCounter)].value+" -- ID -- " + str(DPID.text) + "\r\n")
                                ws['B'+ str(tempCounter)] = DPID.text   #Guarda el valor en el excel./Saves the value on the excel file
                        
                        tempCounter = CounterFilasExcel
                        #Obtains Datapointer identifier of the item
                        #Obtiene el Datapointer-identifier del item.
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                CounterFilasExcel += 1          #Aqui se aumenta el CounterFilasExcel para que se respeten las filas./Counterfilasexcel is updated to keep order of lines
                                logProgram.write("                      "+ws['A'+ str(CounterFilasExcel)].value+" -- DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                ws['M'+ str(CounterFilasExcel)] = DFID.text

                #Cuando la sesion es Reprog./Reprog session
                if(sessionN.text == 'Reprog'):

                        tempCounter = CounterFilasExcel
                        
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel
                        
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                tempCounter += 1
                                logProgram.write("                      "+ws['A'+ str(tempCounter)].value+" -- ID -- " + str(DPID.text) + "\r\n")
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter=CounterFilasExcel
                        
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                CounterFilasExcel += 1
                                logProgram.write("                      "+ws['A'+ str(CounterFilasExcel)].value+" -- DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['I'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es Reprog./Use case Reprog marked on the excel
                #Delivery state session
                #Cuando la sesion es DeliveryState.
                if(sessionN.text == 'DeliveryState'):

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text

                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                tempCounter += 1
                                logProgram.write("                      "+ws['A'+ str(tempCounter)].value+" -- ID -- " + str(DPID.text) + "\r\n")
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                CounterFilasExcel += 1
                                logProgram.write("                      "+ws['A'+ str(CounterFilasExcel)].value+" -- DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['G'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es DeliveryState./Use case DeliveryState marked on the excel

                #Cuando es la sesion es ResetToDeliveryState./Resettodeliverystate marked on the excel
                if(sessionN.text == 'ResetToDeliveryState'):

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                tempCounter += 1
                                logProgram.write("                      "+ws['A'+ str(tempCounter)].value+" -- ID -- " + str(DPID.text) + "\r\n")
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                CounterFilasExcel += 1
                                logProgram.write("                      "+ws['A'+ str(CounterFilasExcel)].value+" -- DESIRED TYPE -- " + str(DFID.text) + "\r\n")
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['H'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es ReturnToDeliveryState./ResetToDeliveryState. marked on the excel
    
    #####################################################
    #Para checar si se repite algun NVM Item.
    #Check if NVM items are repeated
    for i in range(12,CounterAll+12):
        #Agarra cada fila y las compara con las demas./Takes each line and compares it with the rest
        temp = ws['A' + str(i)]
        for j in range(CounterAll+12, CounterFilasExcel+1):
            #Aqui se toma el siguiente en la fila y se checa cada elemento siguiente.//Takes the next the next in line  and checks each element
            temp2 = ws['A' + str(j)]
            #Si son iguales./Items are equal
            if(temp.value == temp2.value):
                #Checa el USE CASES de cada uno./Checks use cases of each element
                temp3 = ws['G' + str(j)]
                if(temp3.value == "X"):
                    #Marca el use case del que se repite ./Marks the use case of the element repeated
                    ws['G'+str(i)]="X"
                temp3 = ws['H'+str(j)]
                if(temp3.value=="X"): 
                    ws['H'+str(i)]="X"
                temp3 = ws['I'+str(j)]
                if(temp3.value=="X"): 
                    ws['I'+str(i)]="X"
                DeleteRepeat+=1

    for k in range(CounterAll+12,CounterFilasExcel+1):
        #Borra la fila que se repite./Erase repeated line
        ws.delete_rows( CounterAll+12,1)

    CounterFilasExcel-=DeleteRepeat   

    logProgram.write("\r\nRepeated Datapointers deleted.\r\n")

    ######################################################
    #Agrega comments.
    #Adds comments
    for datablock in root.iter('DATABLOCK'):
        #Busca en los datablock que nombre tiene./Looks on the the datablocks which name it has
        DBN = datablock.find('DATABLOCK-NAME')
        #Para recorrer el excel./To move through the excel
        for i in range(12, CounterFilasExcel+1):
            #Compara el valor que tiene la celda de excel con el datablock name./Compares the value that the cell has with the datablock name
            if((DBN.text == ((ws['A' + str(i)].value) + '__Metadata')) or (DBN.text==(ws['A' + str(i)].value))):
                #Hay varios DATA por datablock, pero el que se ocupa es el que dice comment/Looks around the data nodes for the comment
                for DPN in datablock.iter('DATA'):
                    temp=DPN.text
                    temp=temp[:11]
                    if(temp=="description"):
                        logProgram.write("DATAPOINTER-NAME -- " + DBN.text + "\r\n")
                        logProgram.write("       COMMENT -- " + str(DPN.text) + "\r\n")
                        #Se copia la descripcion a la columna comment./Copies the comment to the comment column
                        ws['O' + str(i)] = DPN.text

    logProgram.write("\r\nComments added to their corresponding data pointer.\r\n")


    #Si un reporte previo es agregado y se tiene habilitada la opcion /If a previous report is loaded 
    if(archivoReportePrevioCargado == 1 and estadoCheckButton == 1):
        logFile = open(rutaArchivoCNT + "/logFile" + str(BBNumber) + ".txt","w+")
        logProgram.write("\r\nPrevious report added.\r\n")
        sheet1=wb.worksheets[0]
        #Cuenta las filas maximas que tiene el archivo original./ counts the max number of lines that the original file has
        newCounterFilasExcel=sheet1.max_row+1
        wb2=load_workbook(rutaReportePrevio)
        sheet2=wb2.worksheets[0]
        ws2=wb2.active
        #Cuenta cuantos elementos tiene el archivo previo seleccionado./Count how many elements the previous file has
        row_count = sheet2.max_row
        logFile.write("Changes between elements from reports loaded:\r\n\r\n")
        for i in range(12, newCounterFilasExcel):
            for j in range(12, row_count+1):
                if(ws['A'+str(i)].value==ws2['A'+str(j)].value):
                    logFile.write("FROM                     \t" + str(ws['A'+str(i)].value) + "              \tNOT CHANGED               \tCHANGED\r\n")

                    #ID Number./ID Number
                    if(ws['B'+str(i)].value==ws2['B'+str(j)].value):
                        ws['B'+str(i)]=ws2['B'+str(j)].value
                        logFile.write("ID NUMBER                \t"+str(ws['B'+str(i)].value)+"                      \t"+str(ws['B'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("ID NUMBER                \t"+str(ws['B'+str(i)].value)+"             \t                                       \t"+str(ws2['B'+str(j)].value)+"\r\n")
                        ws['B'+str(i)]=ws2['B'+str(j)].value

                    #cr-p./CR-P
                    if(ws['C'+str(i)].value==ws2['C'+str(j)].value):
                        ws['C'+str(i)]=ws2['C'+str(j)].value
                        logFile.write("CR-P                     \t"+str(ws['C'+str(i)].value)+"                     \t"+str(ws['C'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("CR-P                     \t"+str(ws['C'+str(i)].value)+"                \t                                       \t" +str(ws2['C'+str(j)].value) + "\r\n")
                        ws['C'+str(i)]=ws2['C'+str(j)].value

                    #CRP delivery state.
                    if(ws['D'+str(i)].value==ws2['D'+str(j)].value):
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                        logFile.write("CRP DELIVERY STATE       \tX                           \tX\r\n")
                    else:
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                        logFile.write("CRP DELIVERY STATE       \tX                     \t                                       \tX\r\n")

                    #CRP reset delivery state.
                    if(ws['E'+str(i)].value==ws2['E'+str(j)].value):
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                        logFile.write("CRP RESET DELIVERY STATE       \tX                           \tX\r\n")
                    else:
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                        logFile.write("CRP RESET DELIVERY STATE       \tX                     \t                                       \tX\r\n")
                    
                    #CRP reprog.
                    if(ws['F'+str(i)].value==ws2['F'+str(j)].value):
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                        logFile.write("CRP REPROG               \tX                           \tX\r\n")
                    else:
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                        logFile.write("CRP REPROG               \tX                     \t                                       \tX\r\n")
                    
                    #Expected delivery state.
                    if(ws['J'+str(i)].value==ws2['J'+str(j)].value):
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                        logFile.write("EXPECTED DELIVERY STATE       \tX                           \tX\r\n")
                    else:
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                        logFile.write("EXPECTED DELIVERY STATE       \tX                     \t                                       \tX\r\n")
                    
                    #Expected reset delivery state.
                    if(ws['K'+str(i)].value==ws2['K'+str(j)].value):
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                        logFile.write("EXPECTED RESET DELIVERY STATE\tX                           \tX\r\n")
                    else:
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                        logFile.write("EXPECTED RESET DELIVERY STATE\tX                     \t                                       \tX\r\n")
                    
                    #Expected reprog.
                    if(ws['L'+str(i)].value==ws2['L'+str(j)].value):
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                        logFile.write("EXPECTED REPROG              \tX                           \tX\r\n")
                    else:
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                        logFile.write("EXPECTED REPROG              \tX                     \t                                       \tX\r\n")

                    #Desired type.
                    if(ws['M'+str(i)].value==ws2['M'+str(j)].value):
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                        logFile.write("DESIRED TYPE                 \t"+str(ws['M'+str(i)].value)+"                    \t"+str(ws2['M'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("DESIRED TYPE                 \t"+str(ws['M'+str(i)].value)+"              \t                                    \t"+str(ws2['M'+str(j)].value)+"\r\n")
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                        
                    
                    #Desired data.
                    if(ws['N'+str(i)].value==ws2['N'+str(j)].value):
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                        logFile.write("DESIRED DATA                 \t"+str(ws['N'+str(i)].value)+"                     \t"+str(ws2['N'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("DESIRED DATA                 \t"+str(ws['N'+str(i)].value)+"              \t                                      \t"+str(ws2['N'+str(j)].value)+"\r\n")
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                        
                    
                    #Comment.
                    if(ws['O'+str(i)].value==ws2['O'+str(j)].value):
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                        logFile.write("COMMENT                      \t"+str(ws['O'+str(i)].value)+"                             \t"+str(ws2['O'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("COMMENT                      \t"+str(ws['O'+str(i)].value)+"             \t                                            \t"+str(ws2['O'+str(j)].value)+"\r\n")
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                        
                    #Rating.
                    if(ws['P'+str(i)].value==ws2['P'+str(j)].value):
                        ws['P'+str(i)]=ws2['P'+str(j)].value
                        logFile.write("RATING                       \t"+str(ws['P'+str(i)].value)+"                         \t"+str(ws2['P'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("RATING                       \t"+str(ws['P'+str(i)].value)+"             \t                                       \t"+str(ws2['P'+str(j)].value)+"\r\n")
                        ws['P'+str(i)]=ws2['P'+str(j)].value                    
                    
                    #Rated by.
                    if(ws['Q'+str(i)].value==ws2['Q'+str(j)].value):
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                        logFile.write("RATED BY                     \t"+str(ws['Q'+str(i)].value)+"                         \t"+str(ws2['Q'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("RATED BY                     \t"+str(ws['Q'+str(i)].value)+"             \t                                      \t"+str(ws2['Q'+str(j)].value)+"\r\n")
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                    
                    #Comments.
                    if(ws['R'+str(i)].value==ws2['R'+str(j)].value):
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                        logFile.write("COMMENTS                     \t"+str(ws['R'+str(i)].value)+"                      \t"+str(ws2['R'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("COMMENTS                     \t"+str(ws['R'+str(i)].value)+"              \t                                      \t"+str(ws2['R'+str(j)].value)+"\r\n")
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                    
                    #Reference comments from GA.
                    if(ws['S'+str(i)].value==ws2['S'+str(j)].value):
                        ws['S'+str(i)]=ws2['S'+str(j)].value
                        logFile.write("REFERENCE COMMENTS FROM GA   \t"+str(ws['S'+str(i)].value)+"     \t"+str(ws['S'+str(i)].value)+"\r\n")
                    else:
                        logFile.write("REFERENCE COMMENTS FROM GA   \t"+str(ws['S'+str(i)].value)+"                \t                                      \t"+str(ws2['S'+str(j)].value)+"\r\n")
                        ws['S'+str(i)]=ws2['S'+str(j)].value
        
        logFile.write("\r\nNew Elements added to the report:\r\n\r\n")

        for i in range(12, newCounterFilasExcel):
            for j in range(12, row_count+1):
                if(ws['A'+str(i)].value==ws2['A'+str(j)].value):
                    NoRepeat=1
            if(NoRepeat==0):
                logFile.write(str(ws['A'+str(i)].value) + "\r\n")
            NoRepeat=0

        logFile.write("\r\nElements erased from previous report:\r\n\r\n")

        for i in range(12, newCounterFilasExcel):
            for j in range(12, row_count+1):
                if(ws2['A'+str(i)].value==ws['A'+str(j)].value):
                    NoRepeat=1
            if(NoRepeat==0):
                logFile.write(str(ws2['A'+str(i)].value) + "\r\n")
            NoRepeat=0

        logFile.close()
    
    #Guarda los cambios./Save changes
    wb.save(rutaArchivoCNT + "/fillexcel.xlsx")

    #Para ordenar por ID number. Selecciona el archivo excel./Order by ID number, Selects excel file
    excel_file = rutaArchivoCNT + "/fillexcel.xlsx"
    #Leer el archivo./Reads file
    movies = pd.read_excel(excel_file, skiprows=10)
    #Ordena por numero de ID./Ordered  by ID Number
    sorted_by_number = movies.sort_values(by='ID number',ascending=True)
    #Guardar./Save changes
    sorted_by_number.to_excel(excel_file,index=False)

    #Se crea una copia del Template para poder copiar los datos ordenados al archivo que se generara al final./Creates a copy from the template to copy the ordered elements to the final file
    shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Carga el archivo Excel anteriormente generado./Loads Previously generated excel file
    wb1 = load_workbook(filename = rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    ws1=wb1.active

    #Carga el archivo con los datos ordenados./Loads the ordered elements
    wb=load_workbook(filename = rutaArchivoCNT +  "/fillexcel.xlsx")
    sheet=wb.worksheets[0]
    
    #Para que no escriba en espacios vacios./To avoid empty lines
    row_count = sheet.max_row
    ws=wb.active

    #Desde aqui toma los datos del fill excel creado/Take elements from fill excel from the 2nd line
    j=2

    #Los datos los pega ordenados en el archivo excel que es copia del template./Ordered Elements copied to the final file
    #No escribe en espacios vacios./avoids empty spaces
    for i in range(12, 12+row_count-1):
        ws1['A'+str(i)]=ws['A'+str(j)].value
        ws1['B'+str(i)]=ws['B'+str(j)].value
        ws1['C'+str(i)]=ws['C'+str(j)].value
        ws1['D'+str(i)]=ws['D'+str(j)].value
        ws1['E'+str(i)]=ws['E'+str(j)].value
        ws1['F'+str(i)]=ws['F'+str(j)].value
        ws1['G'+str(i)]=ws['G'+str(j)].value
        ws1['H'+str(i)]=ws['H'+str(j)].value
        ws1['I'+str(i)]=ws['I'+str(j)].value
        ws1['J'+str(i)]=ws['J'+str(j)].value
        ws1['K'+str(i)]=ws['K'+str(j)].value
        ws1['L'+str(i)]=ws['L'+str(j)].value
        ws1['M'+str(i)]=ws['M'+str(j)].value
        ws1['N'+str(i)]=ws['N'+str(j)].value
        ws1['O'+str(i)]=ws['O'+str(j)].value
        ws1['P'+str(i)]=ws['P'+str(j)].value
        ws1['Q'+str(i)]=ws['Q'+str(j)].value
        ws1['R'+str(i)]=ws['R'+str(j)].value
        ws1['S'+str(i)]=ws['S'+str(j)].value
        j+=1

    logProgram.write("\r\nExcel file Data Pointers sorted by NVM ID number.\r\n\r\n")
    #Assigns values of BBNumber, baseline, manager and project name tu their respective cells
    #Asigna los valores de BBNumber, Baseline, Encargado y nombre del proyecto a sus respectivas celdas.
    logProgram.write("BBNumber -- " + str(BBNumber) + " added to Excel file.\r\n\r\n")
    logProgram.write("Baseline -- " + str(Baseline) + " added to Excel file.\r\n\r\n")
    logProgram.write("PROJECT MANAGER -- " + PN.text + " added to Excel file.\r\n\r\n")
    logProgram.write("PROJECT DESCRIPTION -- " + PD.text + " added to Excel file.\r\n\r\n")
    ws1['B3']=BBNumber
    ws1['B4']=Baseline
    ws1['B5']=PN.text
    ws1['B2']=PD.text

    #Guarda el archivo./Saves file
    wb1.save(rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Borra el archivo que tiene los datos ordenados./Erase fillexcel file
    os.remove(rutaArchivoCNT + "/fillexcel.xlsx")

    #Close log.
    logProgram.write("END OF LOG FILE\r\n")
    logProgram.close()

def main():
                
    #Titulo de la ventana./Title of the window
    root.title("Bosch") 

    #Configurar paneles./Configure panels
    panelElements.grid(column=0, row=0, sticky=(N, S, E, W))
    panelImage.grid(column=0, row=1, columnspan=2, rowspan=6, sticky=(N, S, E, W))

    #Configurar elementos (botones, etiqueta, imagen, etc)./Configure elements
    image.pack(side = "bottom", fill = "both", expand = "yes")
    label.grid(column=0, row=0, columnspan=4, sticky=(N, W))
    button_CNT.grid(column=3, row=3)
    enable_button.grid(column=3, row=4)
    button_PreviousReport.grid(column=3, row=5)
    button_GenerateReport.grid(column=3, row=6)
        
    #Ocultar botones innecesarios./Hide unnecesary buttons
    button_PreviousReport.grid_remove()
    enable_button.grid_remove()
    button_GenerateReport.grid_remove()

    #Fondos y colores./Background and colors
    style = ttk.Style(root)
    style.configure('TLabel', background='white')       #Background y foreground de la etiqueta./Background and foreground of label
    style.configure('TFrame', background='white')       #Background y foreground del Frame./Background and foreground of Frame

    #Estilo de los elementos./Elements style
    button_style1.configure("button_style1.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'), background = "black", foreground = 'green')
    button_style2.configure("button_style2.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'))
        
    button_CNT.configure(style='button_style1.TButton')
    button_PreviousReport.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style2.TButton')

    #Comenzar con el boton de reporte previo deshabilitado./disabled previous report button
    estadoCheckButton = 0
    button_PreviousReport.state(["disabled"])

    root.config(menu=menubar)
    toolBar_Init = Menu(menubar)
    toolBar_About = Menu(menubar)
    toolBar_Init.add_command(label="New", command=newProgram)
    toolBar_Init.add_command(label="Exit", command=exitProgram)
    toolBar_About.add_command(label="About", command=aboutProgram)
    toolBar_About.add_command(label="Help", command=helpProgram)
    menubar.add_cascade(label="File", menu=toolBar_Init)
    menubar.add_cascade(label="Program", menu=toolBar_About)

    #Comenzar proceso./start process
    root.mainloop()

#################################
if __name__== "__main__":
    main()
