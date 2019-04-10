##################
#### Librerias ###
##################
import os
import sys
import tkinter as tk
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as e

from os import listdir
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter import *

try:
        import xlrd
except ImportError:
        os.system('python -m pip install xlrd')

try:
        import xlwt
except ImportError:
        os.system('python -m pip install xlwt')

try:
        from pathlib import Path
except ImportError:
        os.system('python -m pip install pathlib')

try:
        import shutil
except ImportError:
        os.system('python -m pip install pytest-shutil')

try:
        from PIL import ImageTk, Image
except ImportError:
        os.system('python -m pip install pillow')

try:
        from PIL import ImageTk, Image
except ImportError:
        os.system('python -m pip install pillow')

try:
        from openpyxl import load_workbook
except ImportError:
        os.system('python -m pip install openpyxl')

try:
        import pandas as pd
except ImportError:
        os.system('python -m pip install pandas')

try:
        import win32com.client
except ImportError:
        os.system('python -m pip install pypiwin32')
        os.system('python -m pip install pywin32')

###########################
#### Variables globales ###
###########################
rutaCNT = "0"
rutaArchivoCNT = "0"
archivoCNTCargado = 0
rutaReportePrevio = "0"
archivoReportePrevioCargado = 0
BBNumber = 0
Baseline = 0

estadoCheckButton = 0

################
#### Eventos ###
################
def newProgram():

        global rutaCNT
        global rutaArchivoCNT
        global archivoCNTCargado
        global rutaReportePrevio
        global archivoReportePrevioCargado
        global estadoCheckButton

        #Ocultar botones innecesarios.
        button_PreviousReport.grid_remove()
        enable_button.grid_remove()
        button_GenerateReport.grid_remove()

        button_CNT.configure(style='button_style1.TButton')
        enable_button.deselect()
        button_PreviousReport.configure(style='button_style2.TButton')
        button_GenerateReport.configure(style='button_style2.TButton')

        #Comenzar con el boton de reporte previo deshabilitado.
        estadoCheckButton = 0
        button_PreviousReport.state(["disabled"])

        rutaCNT = "0"
        archivoCNTCargado = 0
        rutaReportePrevio = "0"
        archivoReportePrevioCargado = 0

def exitProgram():
        
        #Preguntar al usuario si desea salir del programa.
        salir = messagebox.askyesno(message="Do you want to close the program?", title="Close program")
        if(salir == 1):
                sys.exit(0)

def aboutProgram():
        
        messagebox.showinfo("About EEPROM report generator", "This software has been released by Julio Cesar Almada Fuerte and Ruben Barajas Curiel")

def helpProgram():
        
        messagebox.showinfo("Help", "Visit the following link to get more information about this software")

def cntButton():

    global rutaCNT
    global archivoCNTCargado

    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening CNT file")
    rutaCNT = filedialog.askopenfilename(filetypes = (("All CNT files","*.cnt"),("All files","*.*")))

    #Verificar que sea archivo CNT o que se haya agregado un archivo.
    if(rutaCNT.find(".cnt") == -1):         #No se selecciono un archivo CNT
        archivoCNTCargado = 0
    elif(rutaCNT == ""):                    #Se dio al boton cancelar.
        archivoCNTCargado = 0
    else:                                   #Se selecciono el archivo correctamente.
        archivoCNTCargado = 1
        button_CNT.configure(style='button_style2.TButton')
        enable_button.grid()                    #Ahora se puede mostrar el checkbuton para habilitar el boton de reporte previo.
        button_PreviousReport.grid()    #Ahora se puede mostrar el boton de reporte previo como opcional.
        button_GenerateReport.grid()    #Ahora se puede mostrar el boton de generar reporte.
        button_GenerateReport.configure(style='button_style1.TButton')

def enableButtonRP():

        global estadoCheckButton

        estadoCheckButton = estadoCheckButton ^ 1

        #Cambiar estado del boton de reporte previo dependiendo del CheckButton.
        if(estadoCheckButton == 0):
                button_PreviousReport.state(["disabled"])
        else:
                button_PreviousReport.state(["!disabled"])

def previousReport():

    global rutaReportePrevio
    global archivoReportePrevioCargado

    #Abrir Dialog Box para buscar el archivo.
    label.configure(text="Opening previous report file")
    rutaReportePrevio = filedialog.askopenfilename(filetypes = (("All Excel files","*.xlsx"),("All files","*.*")))

    #Verificar que sea archivo XLSX o que se haya agregado un archivo.
    if(rutaReportePrevio.find(".xlsx") == -1):       #No se selecciono un archivo XLXS
        archivoReportePrevioCargado = 0
    elif(rutaReportePrevio== ""):                    #Se dio al boton cancelar.
        archivoReportePrevioCargado = 0
    else:                                   #Se selecciono el archivo correctamente.
        archivoReportePrevioCargado = 1

def createReport():

    global archivoCNTCargado
    global archivoReportePrevioCargado
    global rutaCNT
    global rutaArchivoCNT
    global rutaReportePrevio
    global BBNumber
    global Baseline

    label.configure(text="Creating report")

    #Garantizar que se haya seleccionado un archivo CNT.
    if(archivoCNTCargado == 1):                 
        archivoCNTCargado = 0

        #Extraer BBNumber y Baseline.
        BBNumber = rutaCNT.split('_')[3]
        Baseline = rutaCNT.split('_')[4]

        #Crear archivo Excel.
        rutaArchivoCNT = os.path.dirname(rutaCNT)
        shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/fillexcel.xlsx")

        #Rellena Excel
        fillExcel()

        messagebox.showinfo("Report created", "Report created successfully")

        button_CNT.configure(style='button_style1.TButton')
        enable_button.grid_remove()                                                                             #Esconder boton para habilitar reporte previo.
        button_PreviousReport.grid_remove()                                                                     #Esconder boton de reporte previo.
        button_GenerateReport.grid_remove()                                                                     #Esconder boton de generar reporte.
        button_GenerateReport.configure(style='button_style2.TButton')
    else:
        messagebox.showerror("Error", "Not .cnt file selected")

    label.configure(text="EEPROM report generator")

################
#### Objetos ###
################
root = Tk()

rutaActual = os.getcwd()
img = ImageTk.PhotoImage(Image.open(rutaActual + "/bosch.png"))

panelElements = ttk.Frame(root, padding=(3,3,12,12))
panelImage = ttk.Frame(panelElements, borderwidth=5, relief="sunken", width=200, height=200)

label = ttk.Label(panelElements, text="Bosch EEPROM generator", font=("Tahoma", 25, 'bold'))
button_CNT = ttk.Button(panelElements, text="Select CNT file", style="TButton", command=cntButton)
button_PreviousReport = ttk.Button(panelElements, text="Select previous report", style="TButton", command=previousReport)
button_GenerateReport = ttk.Button(panelElements, text="Generate report", style="TButton", command=createReport)
enable_button = Checkbutton(panelElements, text="Enable previous report button", onvalue=1,offvalue=0, command=enableButtonRP)
image = tk.Label(panelImage, image = img)
menubar = Menu(panelImage)

button_style1 = ttk.Style()
button_style2 = ttk.Style()

##################
#### Funciones ###
##################
def fillExcel():

    global rutaArchivoCNT

    #Open log file.
    logProgram = open(rutaArchivoCNT + "/logProgram.txt","w+")

    #Carga el archivo Excel anteriormente generado.
    wb = load_workbook(filename = rutaArchivoCNT + "/fillexcel.xlsx")
    ws = wb.active
    
    #Lee archivo XML.
    tree = ET.parse(rutaCNT)
    logProgram.write("CNT file read.\r\n\r\n")

    #Obtiene el root del XML.
    root = tree.getroot()
    
    #Counter para ir agregando elementos en excel.
    CounterFilasExcel = 11

    ######################################################
    #Guarda el nombre del proyecto.
    for project in root.iter('PROJECT-INFO'):
        logProgram.write("PROJECT-INFO -- " + str(project) +"\r\n")
        PD = project.find('PROJECT-DESC')
        if PD is not None:
                logProgram.write("      PROJECT-DESC -- " + str(PD) + "\r\n")

    ######################################################
    #Guarda el nombre del responsable.
    for info in root.iter('RESPONSIBLE'):
        logProgram.write("RESPONSIBLE -- " + str(info) + "\r\n")
        PN = info.find('PERSON-NAME')
        if PN is not None:
                logProgram.write("      PERSON-NAME -- " + str(PN) + "\r\n")

    ######################################################
    #Busca el nodo sesion en todo el arbol.
    for session in root.iter('SESSION'):
        logProgram.write("SESSION -- " + str(session) + "\r\n")

        #Busca en los tipos de sesiones que nombre tiene.
        sessionN = session.find('SESSION-NAME')
        if sessionN is not None:
                logProgram.write("      SESSION-NAME -- " + str(sessionN) + "\r\n")

                #Cuando la sesion es ALL.
                if(sessionN.text == '__ALL__'):
                        logProgram.write("              __ALL__ -- " + str(sessionN.text) + "\r\n")

                        #Para no alterar el orden de las filas del excel.
                        tempCounter=CounterFilasExcel
                        
                        #Obtiene el Datapointer-name del item.
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text    #Guarda el valor en el excel.
                        
                        tempCounter = CounterFilasExcel
                        
                        #Obtiene el Datapointer-ident del item.
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-IDENT -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text   #Guarda el valor en el excel.
                        
                        tempCounter = CounterFilasExcel
                        
                        #Obtiene el Datapointer-identifier del item.
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-IDENTIFIER -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1          #Aqui se aumenta el CounterFilasExcel para que se respeten las filas.
                                ws['M'+ str(CounterFilasExcel)] = DFID.text

                #Cuando la sesion es Reprog.
                if(sessionN.text == 'Reprog'):
                        logProgram.write("              Reprog -- " + str(sessionN.text) + "\r\n")

                        tempCounter = CounterFilasExcel
                        
                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel
                        
                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-IDENT -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter=CounterFilasExcel
                        
                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-IDENTIFIER -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['I'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es Reprog.

                #Cuando la sesion es DeliveryState.
                if(sessionN.text == 'DeliveryState'):
                        logProgram.write("              DeliveryState -- " + str(sessionN.text) + "\r\n")

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text

                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-IDENT -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-IDENTIFIER -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['G'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es DeliveryState.

                #Cuando es la sesion es ResetToDeliveryState.
                if(sessionN.text == 'ResetToDeliveryState'):
                        logProgram.write("              ResetToDeliveryState -- " + str(sessionN.text) + "\r\n")

                        tempCounter = CounterFilasExcel

                        for DPN in session.iter('DATAPOINTER-NAME'):
                                logProgram.write("                      DATAPOINTER-NAME -- " + str(DPN.text) + "\r\n")
                                tempCounter += 1
                                ws['A'+ str(tempCounter)] = DPN.text
                        
                        tempCounter = CounterFilasExcel

                        for DPID in session.iter('DATAPOINTER-IDENT'):
                                logProgram.write("                      DATAPOINTER-IDENT -- " + str(DPID.text) + "\r\n")
                                tempCounter += 1
                                ws['B'+ str(tempCounter)] = DPID.text
                        
                        tempCounter = CounterFilasExcel

                        for DFID in session.iter('DATAFORMAT-IDENTIFIER'):
                                logProgram.write("                      DATAPOINTER-IDENTIFIER -- " + str(DFID.text) + "\r\n")
                                CounterFilasExcel += 1
                                ws['M'+ str(CounterFilasExcel)] = DFID.text
                                ws['H'+ str(CounterFilasExcel)] = "X"           #Marca que el use case de que es ReturnToDeliveryState.

    ######################################################
    #Agrega comments.
    for datablock in root.iter('DATABLOCK'):
        logProgram.write("DATABLOCK -- " + str(datablock) + "\r\n")
        #Busca en los datablock que nombre tiene.
        DBN = datablock.find('DATABLOCK-NAME')
        #Para recorrer el excel.
        for i in range(12, CounterFilasExcel):
            #Compara el valor que tiene la celda de excel con el datablock name.
            if(DBN.text == ((ws['A' + str(i)].value) + '__Metadata')):
                j = 0
                #Hay varios DATA por datablock, pero el que se ocupa es el 6to.
                for DPN in datablock.iter('DATA'):
                    j += 1
                    if(j == 6):
                        logProgram.write("      	DATA -- " + str(DPN.text) + "\r\n")
                        #Se copia la descripcion a la columna comment.
                        ws['O' + str(i)] = DPN.text

    ######################################################
    #Para checar si se repite algun NVM Item.
    for i in range(12,CounterFilasExcel):
        #Agarra cada fila y las compara con las demas.
        temp = ws['A' + str(i)]
        k = i + 1
        for j in range(k, CounterFilasExcel):
            #Aqui se toma el siguiente en la fila y se checa cada elemento siguiente.
            temp2 = ws['A' + str(j)]
            #Si son iguales.
            if(temp.value == temp2.value):
                #Checa el USE CASES de cada uno.
                temp3 = ws['G' + str(j)]
                if(temp3.value == "X"):
                    #Marca el use case del que se repite .
                    ws['G'+str(i)]="X"
                    #Borra la fila que se repite.
                    ws.delete_rows(j,1)
                temp3 = ws['H'+str(j)]
                if(temp3.value=="X"): 
                    ws['H'+str(i)]="X"
                    ws.delete_rows(j,1)
                temp3 = ws['I'+str(j)]
                if(temp3.value=="X"): 
                    ws['I'+str(i)]="X"
                    ws.delete_rows(j,1)

    #Si un reporte previo es agregado.
    if(archivoReportePrevioCargado):
        logProgram.write("\r\nPREVIOUS REPORT ADDED\r\n\r\n")
        sheet1=wb.worksheets[0]
        #Cuenta las filas maximas que tiene el archivo original.
        newCounterFilasExcel=sheet1.max_row
        wb2=load_workbook(rutaReportePrevio)
        sheet2=wb2.worksheets[0]
        ws2=wb2.active
        #Cuenta cuantos elementos tiene el archivo previo seleccionado.
        row_count = sheet2.max_row
        for i in range(12, newCounterFilasExcel):
            for j in range(12, row_count):
                if(ws['A'+str(i)].value==ws2['A'+str(j)].value):
                    
                    #ID Number.
                    if(ws['B'+str(i)].value==ws2['B'+str(j)].value):
                        ws['B'+str(i)]=ws2['B'+str(j)].value
                    else:
                        ws['B'+str(i)]=ws2['B'+str(j)].value

                    #cr-p.
                    if(ws['C'+str(i)].value==ws2['C'+str(j)].value):
                        ws['C'+str(i)]=ws2['C'+str(j)].value
                    else:
                        ws['C'+str(i)]=ws2['C'+str(j)].value
                    
                    #CRP delivery state.
                    if(ws['D'+str(i)].value==ws2['D'+str(j)].value):
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                    else:
                        ws['D'+str(i)]=ws2['D'+str(j)].value
                    
                    #CRP reset delivery state.
                    if(ws['E'+str(i)].value==ws2['E'+str(j)].value):
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                    else:
                        ws['E'+str(i)]=ws2['E'+str(j)].value
                    
                    #CRP reprog.
                    if(ws['F'+str(i)].value==ws2['F'+str(j)].value):
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                    else:
                        ws['F'+str(i)]=ws2['F'+str(j)].value
                    
                    #Expected delivery state.
                    if(ws['J'+str(i)].value==ws2['J'+str(j)].value):
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                    else:
                        ws['J'+str(i)]=ws2['J'+str(j)].value
                    
                    #Expected reset delivery state.
                    if(ws['K'+str(i)].value==ws2['K'+str(j)].value):
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                    else:
                        ws['K'+str(i)]=ws2['K'+str(j)].value
                    
                    #Expected reprog.
                    if(ws['L'+str(i)].value==ws2['L'+str(j)].value):
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                    else:
                        ws['L'+str(i)]=ws2['L'+str(j)].value
                    
                    #Desired type.
                    if(ws['M'+str(i)].value==ws2['M'+str(j)].value):
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                    else:
                        ws['M'+str(i)]=ws2['M'+str(j)].value
                    
                    #Desired data.
                    if(ws['N'+str(i)].value==ws2['N'+str(j)].value):
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                    else:
                        ws['N'+str(i)]=ws2['N'+str(j)].value
                    
                    #Comment.
                    if(ws['O'+str(i)].value==ws2['O'+str(j)].value):
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                    else:
                        ws['O'+str(i)]=ws2['O'+str(j)].value
                    
                    #Rating.
                    if(ws['P'+str(i)].value==ws2['P'+str(j)].value):
                        ws['P'+str(i)]=ws2['P'+str(j)].value
                    else:
                        ws['P'+str(i)]=ws2['P'+str(j)].value
                    
                    #Rated by.
                    if(ws['Q'+str(i)].value==ws2['Q'+str(j)].value):
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                    else:
                        ws['Q'+str(i)]=ws2['Q'+str(j)].value
                    
                    #Comments.
                    if(ws['R'+str(i)].value==ws2['R'+str(j)].value):
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                    else:
                        ws['R'+str(i)]=ws2['R'+str(j)].value
                    
                    #Reference comments from GA.
                    if(ws['S'+str(i)].value==ws2['S'+str(j)].value):
                        ws['S'+str(i)]=ws2['S'+str(j)].value
                    else:
                        ws['S'+str(i)]=ws2['S'+str(j)].value
    
    #Guarda los cambios.
    wb.save(rutaArchivoCNT + "/fillexcel.xlsx")

    #Para ordenar por ID number. Selecciona el archivo excel.
    logProgram.write("\r\nSORTING BY NVM ID NUMBER\r\n\r\n")
    excel_file = rutaArchivoCNT + "/fillexcel.xlsx"
    #Leer el archivo.
    movies = pd.read_excel(excel_file, skiprows=10)
    #Ordena por numero de ID.
    sorted_by_number = movies.sort_values(by='ID number',ascending=True)
    #Guardar.
    sorted_by_number.to_excel(excel_file,index=False)

    #Se crea una copia del Template para poder copiar los datos ordenados al archivo que se generara al final.
    shutil.copy("EEPROM_Container_Review_Template.xlsx", rutaArchivoCNT + "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Carga el archivo Excel anteriormente generado.
    wb1 = load_workbook(filename = rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    ws1=wb1.active

    #Carga el archivo con los datos ordenados.
    wb=load_workbook(filename = rutaArchivoCNT +  "/fillexcel.xlsx")
    sheet=wb.worksheets[0]
    
    #Para que no escriba en espacios vacios.
    row_count = sheet.max_row
    ws=wb.active

    #Desde aqui toma los datos.
    j=2

    #Los datos los pega ordenados en el archivo excel que es copia del template.
    #No escribe en espacios vacios.
    for i in range(12, 12+row_count-1):
        ws1['A'+str(i)]=ws['A'+str(j)].value
        ws1['B'+str(i)]=ws['B'+str(j)].value
        ws1['C'+str(i)]=ws['C'+str(j)].value
        ws1['D'+str(i)]=ws['D'+str(j)].value
        ws1['E'+str(i)]=ws['E'+str(j)].value
        ws1['F'+str(i)]=ws['F'+str(j)].value
        ws1['G'+str(i)]=ws['G'+str(j)].value
        ws1['H'+str(i)]=ws['H'+str(j)].value
        ws1['I'+str(i)]=ws['I'+str(j)].value
        ws1['J'+str(i)]=ws['J'+str(j)].value
        ws1['K'+str(i)]=ws['K'+str(j)].value
        ws1['L'+str(i)]=ws['L'+str(j)].value
        ws1['M'+str(i)]=ws['M'+str(j)].value
        ws1['N'+str(i)]=ws['N'+str(j)].value
        ws1['O'+str(i)]=ws['O'+str(j)].value
        ws1['P'+str(i)]=ws['P'+str(j)].value
        ws1['Q'+str(i)]=ws['Q'+str(j)].value
        ws1['R'+str(i)]=ws['R'+str(j)].value
        ws1['S'+str(i)]=ws['S'+str(j)].value
        j+=1

    #Asigna los valores de BBNumber, Baseline, Encargado y nombre del proyecto a sus respectivas celdas.
    logProgram.write("BBNumber -- " + str(BBNumber) + "\r\n")
    logProgram.write("Baseline -- " + str(Baseline) + "\r\n")
    logProgram.write("PROJECT MANAGER -- " + str(PN.text) + "\r\n\r\n")
    ws1['B3']=BBNumber
    ws1['B4']=Baseline
    ws1['B5']=PN.text

    #Guarda el archivo.
    wb1.save(rutaArchivoCNT +  "/EEPROM_Container_Review_Checkist_GM_iPB_GlobalB_" + BBNumber + ".xlsx")
    
    #Borra el archivo que tiene los datos ordenados.
    os.remove(rutaArchivoCNT + "/fillexcel.xlsx")

    #Close log.
    logProgram.write("END\r\n")
    logProgram.close()

def main():
                
    #Titulo de la ventana.
    root.title("Bosch") 

    #Configurar paneles.
    panelElements.grid(column=0, row=0, sticky=(N, S, E, W))
    panelImage.grid(column=0, row=1, columnspan=2, rowspan=6, sticky=(N, S, E, W))

    #Configurar elementos (botones, etiqueta, imagen, etc).
    image.pack(side = "bottom", fill = "both", expand = "yes")
    label.grid(column=0, row=0, columnspan=4, sticky=(N, W))
    button_CNT.grid(column=3, row=3)
    enable_button.grid(column=3, row=4)
    button_PreviousReport.grid(column=3, row=5)
    button_GenerateReport.grid(column=3, row=6)
        
    #Ocultar botones innecesarios.
    button_PreviousReport.grid_remove()
    enable_button.grid_remove()
    button_GenerateReport.grid_remove()

    #Fondos y colores.
    style = ttk.Style(root)
    style.configure('TLabel', background='white')       #Background y foreground de la etiqueta.
    style.configure('TFrame', background='white')       #Background y foreground del Frame.

    #Estilo de los elementos.
    button_style1.configure("button_style1.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'), background = "black", foreground = 'green')
    button_style2.configure("button_style2.TButton", width = 20, padding=5, font=('Helvetica', 10, 'bold'))
        
    button_CNT.configure(style='button_style1.TButton')
    button_PreviousReport.configure(style='button_style2.TButton')
    button_GenerateReport.configure(style='button_style2.TButton')

    #Comenzar con el boton de reporte previo deshabilitado.
    estadoCheckButton = 0
    button_PreviousReport.state(["disabled"])

    root.config(menu=menubar)
    toolBar_Init = Menu(menubar)
    toolBar_About = Menu(menubar)
    toolBar_Init.add_command(label="New", command=newProgram)
    toolBar_Init.add_command(label="Exit", command=exitProgram)
    toolBar_About.add_command(label="About", command=aboutProgram)
    toolBar_About.add_command(label="Help", command=helpProgram)
    menubar.add_cascade(label="File", menu=toolBar_Init)
    menubar.add_cascade(label="Program", menu=toolBar_About)

    #Comenzar proceso.
    root.mainloop()

#################################
if __name__== "__main__":
    main()
